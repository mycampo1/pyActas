from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink
from collections import Counter

# Crear una instancia de GoogleAuth o cargar las credenciales almacenadas
gauth_drive = GoogleAuth()

# Intenta cargar las credenciales desde credentials.json
gauth_drive.LoadCredentialsFile("credentials.json")

# Si no hay credenciales almacenadas o están vencidas, solicitará la autenticación
if gauth_drive.credentials is None or gauth_drive.credentials.invalid:
    gauth_drive.LocalWebserverAuth()  # Abre una ventana del navegador para autenticación
    gauth_drive.SaveCredentialsFile("credentials.json")  # Guarda las credenciales en credentials.json

# Crear una instancia de GoogleDrive
drive = GoogleDrive(gauth_drive)

# Definición del ID de la carpeta en Google Drive donde deseas buscar los archivos PDF
carpeta_id = '1HqVrM6jNogxmTp4MMLKuhkbWF0FSXVwx'  # Reemplaza 'tu_id_de_carpeta' con el ID correcto

# Carga el archivo Excel local
wb = load_workbook('C:/Users/Marlon Campo/Downloads/DRIVE NUEVO/ACTAS 2023-1/ACTAS 2023-1.xlsx')
sheet = wb.active

# Inicializa una variable para llevar el control de la fila en la columna 18
fila_columna_18 = 2  # Comienza desde la segunda fila

# Crear una lista de actas para contar su ocurrencia
acta_list = []

# Itera a través de las filas y actualiza la URL en el archivo Excel
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
    acta = row[1]  # Suponiendo que la columna de actas está en la segunda columna (índice 1)

    if row[1] is not None:  # Comprueba si row[1] no es None
        acta_list.append(acta)  # Agregar el número de acta a la lista
        # Realiza la búsqueda en Google Drive
        query = f"'{carpeta_id}' in parents and title='{acta}.pdf'"
        file_list = drive.ListFile({'q': query}).GetList()
        
        if len(file_list) == 1:
            pdf_url = file_list[0]['alternateLink']
            # Coloca la URL en la columna 18, en la fila correspondiente, como enlace hipertexto
            celda_url = sheet.cell(row=fila_columna_18, column=18, value=pdf_url)
            celda_url.font = Font(underline='single', color='0563C1')  # Establece el formato de enlace
            celda_url.hyperlink = pdf_url  # Establece el enlace hipertexto
        else:
            # Si no se encuentra el archivo PDF, coloca "N/A" en la columna 18 como texto normal
            sheet.cell(row=fila_columna_18, column=18, value="N/A")
            sheet.cell(row=fila_columna_18, column=18).hyperlink = None  # Elimina cualquier enlace
            sheet.cell(row=fila_columna_18, column=18).font = Font()  # Establece el formato de texto normal
        
        # Incrementa la fila para la próxima URL
        fila_columna_18 += 1
    
    else:
        # Si row[1] es None, significa que se han agotado los datos, por lo que podemos detener el bucle
        break

# Determina la cantidad de cada número de acta
acta_counts = Counter(acta_list)

# Itera nuevamente a través de la hoja para actualizar con 'N/A' para los números de acta repetidos como texto normal
for fila, acta in enumerate(acta_list, start=2):
    if acta_counts[acta] > 1:
        sheet.cell(row=fila, column=18, value="V/A")
        sheet.cell(row=fila, column=18).hyperlink = None  # Elimina cualquier enlace
        sheet.cell(row=fila, column=18).font = Font()  # Establece el formato de texto normal

# Guarda el archivo Excel actualizado
wb.save('ACTAS 2023-1.xlsx')





#wb = load_workbook('C:/Users/Marlon Campo/Downloads/202/ACTAS 2020-0-refurbished.xlsx')  148DxSRv5VXn-YJ5ZQ8orgKUZMSWJTlQP